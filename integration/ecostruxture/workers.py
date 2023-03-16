"""Ecostruxture Workers module"""

import uuid
from abc import abstractmethod
from collections import Counter
from io import BytesIO
from pathlib import Path
from queue import Queue
from typing import Any, Dict, List, Optional, Set

import openpyxl
import polars as pl
from dataclass_factory import Factory, Schema
from expiringdict import ExpiringDict
from google.cloud.storage import Client
from imap_tools import A, MailBox, MailMessageFlags
from pendulum import DateTime

import common.settings as CFG
from common.bucket_helpers import get_missed_standardized_files, require_client
from common.data_representation.standardized.meter import Meter
from common.date_utils import GapDatePeriod, date_range, format_date, parse, truncate
from common.logging import Logger, ThreadPoolExecutorLogger
from integration.base_integration import BaseFetchWorker, BaseStandardizeWorker
from integration.ecostruxture.data_structures import (
    DataFile,
    FetchedFile,
    StandardizedFile,
)
from integration.ecostruxture.exceptions import EmptyDataInterruption

RUN_GAPS_PARALLEL = True
RUN_FETCH_PARALLEL = True
RUN_STANDARDIZE_PARALLEL = True


class GapsDetectionWorker(BaseFetchWorker):
    """Ecostruxture get missed hours worker functionality"""

    __created_by__ = "Ecostruxture Missed Hours Worker"
    __description__ = "Ecostruxture Integration"
    __name__ = "Ecostruxture Missed Hours Worker"
    __max_idle_run_count__ = 5

    __workers_amount__ = 30

    def __init__(  # pylint:disable=super-init-not-called
        self,
        missed_hours_cache: ExpiringDict,
        config: Any,
    ) -> None:
        self._missed_hours_cache = missed_hours_cache
        self._config = config
        self._meters_queue = Queue()
        self._trace_id: str = uuid.uuid4()
        self._logger = Logger(description=self.__description__, trace_id=self._trace_id)
        self._th_logger = ThreadPoolExecutorLogger(
            description=self.__description__, trace_id=self._trace_id
        )
        self._expected_hours: Optional[pl.LazyFrame] = None

    def configure(self, run_time: DateTime) -> None:
        self._run_time = run_time
        self._clear_queue(self._meters_queue)
        self._missed_hours_cache.clear()
        self._expected_hours = GapDatePeriod(
            self._run_time, self._config.gap_regeneration_window - 1
        )
        for mtr_cfg in self._config.meters:
            self._meters_queue.put(mtr_cfg)

    def missed_hours_consumer(
        self,
        storage_client: Client,
        logs: Queue,  # pylint:disable=unused-argument
        worker_idx: str,  # pylint:disable=unused-argument
    ) -> None:
        """Get missed data points"""
        if self._meters_queue.empty():
            self._th_logger.warning("Meters queue is empty.")
            return None
        empty_run_count = 0

        while True:
            if self._meters_queue.empty():
                if empty_run_count == self.__max_idle_run_count__:
                    break
                empty_run_count += 1
                continue

            mtr_cfg = self._meters_queue.get()
            mtr_msd_poll_hrs = get_missed_standardized_files(
                bucket_name=mtr_cfg.standardized.bucket,
                bucket_path=mtr_cfg.standardized.path,
                range_hours=self._config.gap_regeneration_window,
                client=storage_client,
                date_range=self._expected_hours,
            )
            if mtr_msd_poll_hrs:
                for mtr_hr in mtr_msd_poll_hrs:
                    # TODO: @totdo Fix to add ability to use multiple configs per few hours
                    # See Wattime as the example
                    try:
                        self._missed_hours_cache[mtr_hr].put(mtr_cfg)
                    except KeyError:
                        self._missed_hours_cache[mtr_hr] = Queue()
                        self._missed_hours_cache[mtr_hr].put(mtr_cfg)

                self._th_logger.info(
                    f"Found {len(mtr_msd_poll_hrs)} in '{mtr_cfg.meter_name}' "
                    "meter.",
                )
            else:
                self._th_logger.info(f"Meter {mtr_cfg.meter_name} is up to date.")
            self._meters_queue.task_done()

    def run(self, run_time: DateTime) -> None:
        """Run loop entrypoint"""
        self.configure(run_time)
        self._run_consumers(
            [(self.missed_hours_consumer, [require_client()])],
            run_parallel=RUN_GAPS_PARALLEL,
        )

    # TODO: should be removed
    @abstractmethod
    def run_fetch_worker(self, logs: Queue, worker_idx: int) -> None:
        """Run fetch worker"""


class FetchWorker(BaseFetchWorker):
    """Ecostruxture fetch worker functionality"""

    __created_by__ = "Ecostruxture Fetch Connector"
    __description__ = "Ecostruxture Integration"
    __name__ = "Ecostruxture Fetch Worker"

    __mail_root_folder__ = Path("INBOX")
    __mail_processed__ = "ECOSTRUXTURE_PROCESSED"
    __max_pool_size__ = 20

    __sheet_first_value_row__ = 2
    __sheet_date_column__ = 1

    __fetch_file_name_tmpl__ = "{base_file_name}_{idx}"
    __sheet_execess_cols__ = (3, 4)
    __max_idle_run_count__ = 5

    __raw_consumption_col__ = "RawConsumption"
    __raw_meter_hour_col__ = "RawDates"
    __df_timestamp_coll__ = "Timestamp"
    __df_date_times_col__ = "Dates"
    __df_aligned_date_times_col__ = "AlignedDates"
    __df_adjusted_date_times_col__ = "AdjustedDates"
    __df_aligned_adjusted_date_times_col__ = "AlignedAdjustedDates"
    __df_meter_hour_col__ = "MeterHour"
    __df_consumption_col__ = "Consumption"

    __default_meter_date__ = "1970-01-01T00:00"
    __default_raw_consumption = 0

    def __init__(
        self,
        missed_hours: ExpiringDict,
        fetched_files: Queue,
        fetch_update: Queue,
        config: Any,
    ) -> None:
        super().__init__(
            missed_hours=missed_hours,
            fetched_files=fetched_files,
            fetch_update=fetch_update,
            config=config,
        )

        self._missed_hours_queue: ExpiringDict = missed_hours
        self._tmp_task_q: Queue = Queue()

        self._fetched_atachments_q = Queue()

        self._processed_messages_q = Queue()
        self._skipped_messages_q = Queue()

        self._fetch_counter = Counter()

        self._base_name: Optional[str] = None

        self._factory = Factory(
            default_schema=Schema(trim_trailing_underscore=False, skip_internal=False)
        )

        self._allowed_sheet_names: Optional[Set] = None

    def configure(self, run_time: DateTime) -> None:
        super().configure(run_time=run_time)
        self._clear_queue(self._fetched_atachments_q)
        self._clear_queue(self._processed_messages_q)
        self._clear_queue(self._skipped_messages_q)
        self._fetch_counter.clear()
        self._base_name = format_date(self._run_time, CFG.PROCESSING_DATE_FORMAT)
        self._allowed_sheet_names = set(self._config.meters_sheet_mapper.values())

    def _get_mail_connection(self) -> MailBox:
        return MailBox(self._config.host).login(
            self._config.email, self._config.password
        )

    def _create_mailbox_folder(self, mailbox: MailBox, folder: str) -> None:
        is_exists = mailbox.folder.exists(folder)
        if not is_exists:
            mailbox.folder.create(folder)
        else:
            self._th_logger.warning(f"Mailbox folder {folder} alredy exist.")

    @staticmethod
    def _joinmailpath(*args):
        root = Path(args[0])
        return str(root.joinpath(*args[1:])).replace(
            root._flavour.sep, "|"  # pylint: disable=protected-access,no-member
        )

    def _get_mail_processed_folder(self, mailbox: MailBox) -> str:
        processed_folder = self._joinmailpath(
            self.__mail_root_folder__, self.__mail_processed__
        )
        self._create_mailbox_folder(mailbox=mailbox, folder=processed_folder)
        return processed_folder

    def _get_message_query(self) -> A:
        return A(subject=self._config.subject, from_=self._config.sender, seen=False)

    def _move_processed_msgs(self, mailbox: MailBox) -> None:
        if not self._processed_messages_q.empty():
            folder = self._get_mail_processed_folder(mailbox)
            uids = list(self._processed_messages_q.queue)
            mailbox.flag(uids, (MailMessageFlags.SEEN,), True)
            mailbox.move(uids, folder)
            self._th_logger.info(f"Moved '{','.join(uids)}' to the folder {folder}")
            self._clear_queue(self._processed_messages_q)

    def _mark_as_unread_skipped_msgs(self, mailbox: MailBox) -> None:
        if not self._skipped_messages_q.empty():
            uids = list(self._skipped_messages_q.queue)
            mailbox.flag(uids, (MailMessageFlags.SEEN,), False)
            self._th_logger.info(f"Prepared'{','.join(uids)}' to next fetch iteration")
            self._clear_queue(self._skipped_messages_q)

    def filter_mail_attachments(self, attachemnts: List) -> List:
        """Filter mail attachemts to drop excess"""
        return list(
            filter(
                lambda x: x.filename.strip().lower() == self._config.attachment,
                attachemnts,
            )
        )

    def fetch(self) -> None:
        """Get new excel"""
        idle = False
        with self._get_mail_connection() as mailbox:
            query = self._get_message_query()

            msgs = list(mailbox.fetch(query))

            for msg in msgs:
                if idle:
                    self._skipped_messages_q.put(msg.uid)
                    continue

                attachements = self.filter_mail_attachments(msg.attachments)
                qsize = self._fetch_counter["atachements"]
                if qsize + len(attachements) >= self.__max_pool_size__:
                    idle = True
                    self._skipped_messages_q.put(msg.uid)
                    continue

                self._processed_messages_q.put(msg.uid)
                for attachement in attachements:
                    self._fetched_atachments_q.put((msg.uid, attachement))

            self._move_processed_msgs(mailbox)
            self._mark_as_unread_skipped_msgs(mailbox)

            if self._fetched_atachments_q.empty():
                self._logger.warning(
                    "The new messages or related attachements were not found."
                )

    def _read_excel(self, data: bytes):
        excel = openpyxl.reader.excel.load_workbook(data, read_only=False)
        excel.remove_sheet(excel["Summary"])
        return excel

    def _get_meter_hour_form_excel(self, workbook) -> Set[DateTime]:
        sheet = workbook.active
        date = sheet.cell(
            row=self.__sheet_first_value_row__, column=self.__sheet_date_column__
        ).value

        mtr_dt = truncate(parse(date), level="hour")

        day_hours = date_range(
            start_date=mtr_dt.start_of("day"), end_date=mtr_dt.end_of("day")
        )
        return set(day_hours)

    def _move_skipped_to_inbox(self) -> None:
        if not self._skipped_messages_q.empty():
            with self._get_mail_connection() as mailbox:
                uids = list(self._skipped_messages_q.queue)
                mailbox.flag(uids, (MailMessageFlags.SEEN,), False)
                mailbox.move(uids, str(self.__mail_root_folder__))
                self._th_logger.info(
                    f"Moved '{','.join(uids)}' to the folder "
                    f"{self.__mail_root_folder__}"
                )
                self._clear_queue(self._skipped_messages_q)

    def __df_rename_columns(self, data_df: pl.LazyFrame) -> pl.LazyFrame:
        return data_df.lazy().rename(
            {
                "column_0": self.__raw_meter_hour_col__,
                "column_1": self.__raw_consumption_col__,
            }
        )

    @staticmethod
    def __df_set_date_type(
        data_df: pl.LazyFrame, date_col: str, dest_col: str
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            (pl.col(date_col).apply(lambda y: parse(y, tz_info="UTC"))).alias(dest_col)
        )

    def __df_adjust_dates(
        self, data_df: pl.LazyFrame, date_col: str, dest_col: str
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            (pl.col(date_col).apply(self._adjust_meter_date)).alias(dest_col)
        )

    @staticmethod
    def __df_truncate_dates(
        data_df: pl.LazyFrame, date_col: str, dest_col: str, level: str
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            (pl.col(date_col).apply(lambda y: truncate(y, level=level))).alias(dest_col)
        )

    @staticmethod
    def __df_get_timestamp(
        data_df: pl.LazyFrame, date_col: str, dest_col: str
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            (pl.col(date_col).apply(lambda y: int(y.timestamp()))).alias(dest_col)
        )

    @staticmethod
    def __df_get_meter_hour(
        data_df: pl.LazyFrame, date_col: str, dest_col: str
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            (
                pl.col(date_col).apply(
                    lambda y: format_date(y, CFG.PROCESSING_DATE_FORMAT)
                )
            ).alias(dest_col)
        )

    @staticmethod
    def __df_remove_dublicates(data_df: pl.LazyFrame) -> pl.LazyFrame:
        return data_df.lazy().unique()

    @staticmethod
    def __df_remove_nulls(data_df: pl.LazyFrame, columns: List[str]) -> pl.LazyFrame:
        return data_df.lazy().filter(pl.col(columns).is_null().is_not())

    @staticmethod
    def __df_fill_nulls(data_df: pl.LazyFrame, column: str, value: Any) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            pl.col(column).fill_null(
                pl.lit(value),
            ),
        )

    @staticmethod
    def __df_set_type_inline(
        data_df: pl.LazyFrame, column: str, pl_type: Any
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            pl.col(column).cast(pl_type, strict=False).keep_name()
        )

    def _get_data_pl_df(self, data: Any) -> pl.LazyFrame:
        if not data:
            raise EmptyDataInterruption("Recieved Empty Data")

        sheet_data = data.values
        next(sheet_data)[0:]  # pylint:disable=expression-not-assigned
        mtr_df = pl.DataFrame(sheet_data).lazy()

        return (
            mtr_df.pipe(self.__df_rename_columns)
            .pipe(self.__df_remove_dublicates)
            .pipe(
                self.__df_fill_nulls,
                self.__raw_meter_hour_col__,
                self.__default_meter_date__,
            )
            .pipe(self.__df_set_type_inline, self.__raw_consumption_col__, pl.Float64)
            .pipe(
                self.__df_fill_nulls,
                self.__raw_consumption_col__,
                self.__default_raw_consumption,
            )
            .pipe(self.__df_remove_dublicates)
            .pipe(
                self.__df_remove_nulls,
                [self.__raw_consumption_col__, self.__raw_meter_hour_col__],
            )
            .pipe(
                self.__df_set_date_type,
                self.__raw_meter_hour_col__,
                self.__df_date_times_col__,
            )
            .pipe(
                self.__df_adjust_dates,
                self.__df_date_times_col__,
                self.__df_adjusted_date_times_col__,
            )
            .pipe(
                self.__df_truncate_dates,
                self.__df_date_times_col__,
                self.__df_aligned_date_times_col__,
                "hour",
            )
            .pipe(
                self.__df_truncate_dates,
                self.__df_adjusted_date_times_col__,
                self.__df_aligned_adjusted_date_times_col__,
                "hour",
            )
            .pipe(
                self.__df_get_meter_hour,
                self.__df_aligned_adjusted_date_times_col__,
                self.__df_meter_hour_col__,
            )
            .pipe(
                self.__df_get_timestamp,
                self.__df_adjusted_date_times_col__,
                self.__df_timestamp_coll__,
            )
            .sort([self.__df_timestamp_coll__])
        )

    @staticmethod
    def __df_min_inline(
        data_df: pl.LazyFrame, data_col: str, group_col: str, suffix: str = "_min"
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            (pl.col(data_col).min().over(group_col).suffix(suffix))
        )

    @staticmethod
    def __df_max_inline(
        data_df: pl.LazyFrame, data_col: str, group_col: str, suffix: str = "_max"
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            (pl.col(data_col).max().over(group_col).suffix(suffix))
        )

    @staticmethod
    def __df_max_inline(
        data_df: pl.LazyFrame, data_col: str, group_col: str, suffix: str = "_max"
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            (pl.col(data_col).max().over(group_col).suffix(suffix))
        )

    @staticmethod
    def __df_difference_inline(
        data_df: pl.LazyFrame, minuend_col: str, subtrahend_col: str, res_col: str
    ) -> pl.LazyFrame:
        return data_df.lazy().with_columns(
            ((pl.col(minuend_col) - pl.col(subtrahend_col)).alias(res_col))
        )

    def _get_consumption_pl_df(self, data_df: pl.LazyFrame) -> pl.LazyFrame:

        # For debug purpose we need to store all data in df for some period.
        # TODO: @ todo After some probation period shoulbe chaged to
        # return data_df.groupby("MeterHour").agg(
        #   pl.col("RawConsumption").max() - pl.col("RawConsumption").min()
        # ).collect()

        return (
            data_df.pipe(
                self.__df_min_inline,
                self.__raw_consumption_col__,
                self.__df_meter_hour_col__,
            )
            .pipe(
                self.__df_max_inline,
                self.__raw_consumption_col__,
                self.__df_meter_hour_col__,
            )
            .pipe(
                self.__df_difference_inline,
                f"{self.__raw_consumption_col__}_max",
                f"{self.__raw_consumption_col__}_min",
                self.__df_consumption_col__,
            )
        )

    def _get_workbook_idx(self, workbook: Any) -> Dict:
        sheets_idx = {}

        for sheet_name in workbook.sheetnames:
            if sheet_name not in self._allowed_sheet_names:
                self._th_logger.warning(f"Found unexpected sheet '{sheet_name}'. Skip")
                continue
            sheet = workbook.get_sheet_by_name(sheet_name)
            sheet.delete_cols(*self.__sheet_execess_cols__)
            try:
                raw_data_df = self._get_data_pl_df(sheet)
                raw_data_df = self._get_consumption_pl_df(raw_data_df)
            except EmptyDataInterruption as err:
                self._th_logger.error(
                    f"Can not load meter data to a Dataframe due to the err '{err}'"
                )
                continue
            else:
                sheets_idx[sheet_name] = raw_data_df
        return sheets_idx

    def unbundle_fetch_data_consumer(
        self,
        logs: Queue,  # pylint:disable=unused-argument
        worker_idx: str,  # pylint:disable=unused-argument
    ) -> None:
        """Split excel document by sheets for paralelization"""
        if self._fetched_atachments_q.empty():
            self._th_logger.warning("Attachements queue is empty.")
            return None
        empty_run_count = 0
        while True:
            if self._fetched_atachments_q.empty():
                if empty_run_count == self.__max_idle_run_count__:
                    self._move_skipped_to_inbox()
                    break

                empty_run_count += 1
                continue
            msg_uid, attachement = self._fetched_atachments_q.get()

            mtr_wb = self._read_excel(BytesIO(attachement.payload))
            day_hours = self._get_meter_hour_form_excel(mtr_wb)

            hours = set(self._missed_hours_queue.keys()).intersection(day_hours)

            if hours:
                with self._lock:
                    filename = self.__fetch_file_name_tmpl__.format(
                        base_file_name=self._base_name,
                        idx=self._fetch_counter["fetched_id"],
                    )
                    self._fetch_counter["fetched_id"] += 1

                file_info = FetchedFile(
                    file_name=filename,
                    bucket=self._config.extra.raw.bucket,
                    path=self._config.extra.raw.path,
                    body=attachement.payload,
                )
                self._shadow_fetched_files_queue.put(file_info)
                self._add_to_update(file_info, self._fetch_update_file_buffer)
                sheets_idx = self._get_workbook_idx(mtr_wb)

                for mtr_hr in hours:
                    with self._lock:
                        mtr_cfgs = self._missed_hours_queue.get(mtr_hr, None)
                        mtr_cfgs = list(mtr_cfgs.queue) if mtr_cfgs else []
                    for mtr_cfg in mtr_cfgs:
                        mt_name = Path(mtr_cfg.meter_name).stem
                        sheet_name = self._config.meters_sheet_mapper.get(mt_name, "")
                        if not sheet_name or sheet_name not in sheets_idx:
                            self._th_logger.warning(
                                f"Cannot find sheet '{sheet_name}' related "
                                f"to meter '{mt_name}' in gs://"
                                f"{self._config.extra.raw.bucket}/"
                                f"{self._config.extra.raw.path}/{filename}"
                            )
                            continue

                        data_file = DataFile(
                            file_name=filename,
                            bucket=self._config.extra.raw.bucket,
                            path=self._config.extra.raw.path,
                            body=sheets_idx[sheet_name].collect(),
                        )
                        data_file.timestamps.put(mtr_hr)
                        data_file.meters.put(mtr_cfg)
                        self._fetched_files_queue.put(data_file)
            else:
                self._skipped_messages_q.put(msg_uid)
            self._fetched_atachments_q.task_done()

    def run(self, run_time: DateTime) -> None:
        """Run loop entrypoint"""
        self.configure(run_time)
        self.fetch()

        self._run_consumers(
            [
                (self.unbundle_fetch_data_consumer, []),
                (self.save_fetched_files_worker, []),
            ],
            run_parallel=RUN_FETCH_PARALLEL,
        )
        self.finalize_fetch_update_status()
        self._run_consumers(
            [
                (self.save_fetch_status_worker, []),
            ],
            run_parallel=RUN_FETCH_PARALLEL,
        )
        self._logger.info("Fetching has been done.")

    # TODO: should be removed
    @abstractmethod
    def run_fetch_worker(self, logs: Queue, worker_idx: int) -> None:
        """Run fetch worker"""


class StandardizeWorker(BaseStandardizeWorker):
    """IES Mach Standardization Worker."""

    __created_by__ = "IES Mach Connector"
    __description__ = "IES Mach Integration"
    __name__ = "IES Mach Standardize Worker"

    __df_meter_hour_col__ = "MeterHour"
    __df_consumption_col__ = "Consumption"

    def _standardize_electric(
        self, data_df: pl.DataFrame(), mtr_date: DateTime, mtr_cfg: Dict
    ) -> Meter:
        """Standardize electric value"""
        if not bool(len(data_df)):
            raise EmptyDataInterruption(
                f"Recieved Empty Data for the meter '{mtr_cfg.meter_name}'"
            )

        start_date = truncate(mtr_date, level="hour")
        end_date = start_date.add(minutes=59, seconds=59)

        mtr_date = self._adjust_meter_date(start_date)
        mtr_hr = format_date(mtr_date, CFG.PROCESSING_DATE_FORMAT)

        try:
            consumption = data_df.filter(pl.col(self.__df_meter_hour_col__) == mtr_hr)[
                self.__df_consumption_col__
            ][0]
        except pl.exceptions.ComputeError:
            raise EmptyDataInterruption(  # pylint:disable=raise-missing-from
                f"Data Frame of '{mtr_cfg.meter_name}' does not contain data for "
                f"'{mtr_hr}' point"
            )

        def _getter(
            xdata: List[Dict[str, Any]]
        ) -> Optional[str]:  # pylint:disable=unused-argument
            return consumption, start_date, end_date

        return self._standardize_generic([], mtr_cfg, _getter)

    def _standardize(self, raw_file_obj: DataFile) -> List[DataFile]:
        """Standardize the given raw file"""

        # TODO: Redesign worker to process all meter points in parallel
        standardized_files = []
        try:
            raw_data_df = raw_file_obj.body
            if not bool(len(raw_data_df)):
                self._th_logger.error("Recived empty standardize data. Skipping")
                return []
        except EmptyDataInterruption as err:
            self._th_logger.error(
                f"Can not load meter data to a Dataframe due to the err '{err}'"
            )
            return []

        while not raw_file_obj.timestamps.empty():
            mtr_hr = raw_file_obj.timestamps.get()
            while not raw_file_obj.meters.empty():
                mtr_cfg = raw_file_obj.meters.get()
                meter_type = mtr_cfg.type.strip().lower().replace(" ", "_")
                stndrdz_mthd_nm = f"_standardize_{meter_type}"
                stndrdz_func = getattr(self, stndrdz_mthd_nm, "")
                if not callable(stndrdz_func):
                    #  TODO: Add loging here
                    continue

                try:
                    meter = stndrdz_func(
                        data_df=raw_data_df,
                        mtr_date=mtr_hr,
                        mtr_cfg=mtr_cfg,
                    )
                except EmptyDataInterruption as err:
                    self._th_logger.error(
                        f"Cannot standardize meter {mtr_cfg.meter_name}"
                        f" point '{mtr_hr}' due to the error '{err}'"
                    )
                    continue

                standardized_files.append(
                    StandardizedFile(
                        file_name=format_date(mtr_hr, CFG.PROCESSING_DATE_FORMAT),
                        bucket=mtr_cfg.standardized.bucket,
                        path=mtr_cfg.standardized.path,
                        meter=meter,
                        body=meter.as_str(),
                        cfg=mtr_cfg,
                    )
                )
                raw_file_obj.meters.task_done()
            raw_file_obj.timestamps.task_done()
        return standardized_files

    # TODO: @todo Possible candite to be in base class. Or boiler plate code
    def run(self, run_time: DateTime) -> None:
        """Run loop entrypoint"""
        self.configure(run_time)
        self._run_consumers(
            [
                (self.run_standardize_worker, []),
                (self.save_standardized_files_worker, []),
            ],
            run_parallel=RUN_STANDARDIZE_PARALLEL,
        )
        self.finalize_standardize_update_status()
        self._run_consumers(
            [(self.save_standardize_status_worker, [])],
            run_parallel=RUN_STANDARDIZE_PARALLEL,
        )
